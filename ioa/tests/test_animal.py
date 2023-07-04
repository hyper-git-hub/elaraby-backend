import traceback

from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
from django.utils.crypto import get_random_string
from hypernet.enums import DeviceTypeAssignmentEnum
from jsonpickle.compat import unicode
from rest_framework.permissions import AllowAny

from hypernet.entity.utils import single_or_bulk_delete_check_related_objects
from iof.serializers import CMSVehicleReportingSerializer

__author__ = 'nahmed'

from django.db.models.signals import post_save, post_delete, pre_save
import csv
from django.dispatch import receiver
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from hypernet.utils import generic_response, get_customer_from_request, get_module_from_request, response_json
from ioa.utils import *
from reportlab import *


import json
import requests
import logging
import hmac
import base64
from hashlib import sha256
#from .constants import *
from sys import getsizeof
from django.http import HttpResponse
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from hypernet.enums import DeviceTypeEntityEnum


# from asyncio import

@csrf_exempt
@api_view(['GET'])
@exception_handler(generic_response(response_body=ERROR_RESPONSE_BODY, http_status=500))
def animal_state_violation(request):
    from ioa.animal.animal_utils import animal_state_cron
    animal_state_cron()
    response_body = {RESPONSE_MESSAGE: "", RESPONSE_STATUS: STATUS_OK, RESPONSE_DATA: []}

    return generic_response(response_body=response_body, http_status=200)


@csrf_exempt
@api_view(['GET'])
@exception_handler(generic_response(response_body=ERROR_RESPONSE_BODY, http_status=500))
def animal_signal_test(request):
    from hypernet.models import Entity
    from customer.models import Customer
    customer = Customer.objects.filter(id=1)[0]

    entity = Entity()
    entity.name = customer.name + '_cow_' + str(101)
    entity.assignments = 0
    entity.device_name = None
    entity.speed = False
    entity.volume = False
    entity.location = False
    entity.temperature = False
    entity.density = False
    entity.obd2_compliant = False
    entity.lactation_days = randint(0, 180)
    entity.group_id = randint(1011, 1013)
    entity.lactation_status_id = randint(1017, 1019)
    entity.customer = customer
    entity.modified_by_id = 1
    entity.module_id = ModuleEnum.IOA
    entity.status = Options.objects.get(value='Active', key='recordstatus')
    entity.type_id = DeviceTypeEntityEnum.ANIMAL
    entity.age = randint(1, 10)
    entity.weight = randint(100, 300)
    entity.save()
    response_body = {RESPONSE_MESSAGE: "", RESPONSE_STATUS: STATUS_OK, RESPONSE_DATA: []}

    return generic_response(response_body=response_body, http_status=200)


@csrf_exempt
@api_view(['GET'])
@exception_handler(generic_response(response_body=ERROR_RESPONSE_BODY, http_status=500))
def test_logging(request):
    import logging
    logger = logging.getLogger('hypernet')
    logger.debug('Logging DEBUG')
    logger.info('Logging INFO')
    logger.warning('Logging WARN')
    logger.error('Logging ERROR')
    # import hello
    try:
        import _json
    except ImportError as err:
        # logger.exception(err)
        logger.error(err, exc_info=True)
    response_body = {RESPONSE_MESSAGE: "", RESPONSE_STATUS: STATUS_OK, RESPONSE_DATA: []}
    return generic_response(response_body=response_body, http_status=200)


@csrf_exempt
@api_view(['GET'])
@exception_handler(generic_response(response_body=ERROR_RESPONSE_BODY, http_status=500))
def hypernet_search_bar(self):
    from django.db.models.functions import Concat
    from django.contrib.postgres.search import SearchVector
    response_body = {RESPONSE_MESSAGE: "", RESPONSE_STATUS: STATUS_OK, RESPONSE_DATA: []}
    search_q = self.query_params.get('search')
    customer = get_customer_from_request(self, None)
    module_id = get_module_from_request(self, None)
    entity = Entity.objects.filter(customer=customer, module_id=module_id)
    search_results = entity.filter(). \
                         annotate(results=Concat('name', 'lactation_status__label', 'group__label', 'type__name')). \
                         filter(results__icontains=search_q)[:10]
    # for result in search_results:
    #     response_body[RESPONSE_DATA].append(result.animal_details_to_dict())
    response_body[RESPONSE_DATA] = list(search_results.values('id', title=F('name'), entity_type=F('type__name')))
    return generic_response(response_body=response_body, http_status=200)


# TEST CALL TO SAVE LIST OF DICT TO CSV
# ---------------------------------------------------------------------------------------------------------------------
@csrf_exempt
@api_view(['POST'])
@exception_handler(generic_response(response_body=ERROR_RESPONSE_BODY, http_status=500))
def data_to_csv(request):
    from email_manager.email_util import extended_email_with_title
    import csv
    from itertools import zip_longest
    response_body = {RESPONSE_MESSAGE: "Data Added to File", RESPONSE_STATUS: STATUS_OK, RESPONSE_DATA: {}}

    q_set = Entity.objects.get(id=10295)

    flag, ent = single_or_bulk_delete_check_related_objects(q_set)
    response_body[RESPONSE_DATA] = [obj.as_entity_json() for obj in ent]
    # animals = Entity.objects.filter(customer_id=get_customer_from_request(request, None),
    #                                 module_id=ModuleEnum.IOL, type_id=DeviceTypeEntityEnum.TRUCK).values('name',
    #                                                                                                           'customer__name',
    #                                                                                                           'model', )
    # a = Assignment
    # dump(qs=Assignment, outfile_path='Test.csv')
    # export_data = zip_longest(*animals, fillvalue='')
    # with open('Test.csv', 'w', encoding="ISO-8859-1", newline='') as CSV_FILE:
    #     fields = ["name", "customer__name", "model"]
    #     csv_writer = csv.DictWriter(CSV_FILE, fieldnames=fields)
    #     # csv_writer.writerow(("name", "customer__name", "model"))
    #     for i in animals:
    #         csv_writer.writeheader()
    #         csv_writer.writerow(i.name)#, "customer__name":data.customer.name, "model": data.model})
    #         csv_writer.writerow(i.customer.name)
    # from email_manager.email_util import extended_email_with_title
    # csv_writer.writerow(['name', 'customer__name', 'model', 'modified_datetime'])
    # extended_email_with_title(title='activity_performed',
    #                           email_words_dict={})
    # send_sms_util()
    # response_body[RESPONSE_DATA] = cron_job_IOA_alerts()

    # import requests
    # email = request.data["email"]
    # password = request.data["password"]
    # data = {'email': email, 'password': password}
    # r = requests.post(url='http://188.166.226.185/api/users/login/', data=data)
    # result = r.content
    # jso = json.loads(result.decode('utf'))
    # response_body[RESPONSE_DATA] = jso["response"]
    return generic_response(response_body=response_body, http_status=200)


def dump(qs, outfile_path='Test.csv'):
    """
    Takes in a Django queryset and spits out a CSV file.

    Usage::

        >> from utils import dump2csv
        >> from dummy_app.models import *
        >> qs = DummyModel.objects.all()
        >> dump2csv.dump(qs, './data/dump.csv')

    Based on a snippet by zbyte64::

        http://www.djangosnippets.org/snippets/790/

    """
    model = qs
    writer = csv.writer(open(outfile_path, 'w'))
    headers = []

    field = [f for f in model._meta.fields if f.name != 'id' and f.name != 'child']

    for i in field:
        headers.append(i.name)
    writer.writerow(headers)
    for obj in qs.objects.all():
        row = []
        for field in headers:
            val = getattr(obj, field)
            if callable(val):
                val = val()
            # if type(val) == unicode:
            #     val = val.encode("utf-8")
            row.append(val)
        writer.writerow(row)

#TODO Uncoment
'''
@receiver(post_save, sender=CustomerDevice)
def update_proxy_entitymap(sender, instance, **kwargs):
    from hypernet.models import EntityMap
    if instance.type_id == DeviceTypeEntityEnum.IOP_DEVICE:
        pass

    else:
        em, flag = EntityMap.objects.update_or_create(
            device_id=instance.device_id,
            defaults={
                'primary_key': instance.primary_key,
                'device_id':instance.device_id,
                'device_name':instance.device_id,
                'customer':instance.customer.id,
                'module':instance.module.id,
                'type':instance.type_id,
                'iot_hub':instance.connection_string
                    })

    # if instance.type_id == DeviceTypeEntityEnum.ANIMAL:
    #     em.animal_group = instance.group.value
    #     em.lactation_status = instance.lactation_status.value
    #     em.breed = instance.breed.value
    #     em.save()


@receiver(post_delete, sender=CustomerDevice)
def delete_proxy_entitymap(sender, instance, **kwargs):
    from hypernet.models import EntityMap
    try:
        EntityMap.objects.filter(
            type=instance.type_id,
            primary_key=instance.primary_key,
        ).delete()

    except EntityMap.DoesNotExist:
        pass

    return None

'''
# @receiver(pre_save, sender=CustomerDevice)
# def update_proxy_entitymap(sender, instance, **kwargs):
#     from hypernet.models import EntityMap
#     return None

def generate_sas_token(iot_hub, key, device_id, ttl=10):
    """
        Given the device name and primary key, generates a SAS token.
        ttl is the validity time for the token i.e. default is 10 minutes.

    :param key: The primary key, generated when a device is added on IOT hub.
    :param device_name: The device name on IOT hub.
    :param ttl: Time to Live in minutes.
    :return:
    """
    import datetime
    import time
    import urllib.parse as urllib

    url = "/devices/" + device_id
    host = "{0}.azure-devices.net".format(iot_hub)
    encoded_url = urllib.quote(host + url)  # uses safe itself
    print("Encoded url", encoded_url)
    # +10 minutes timestamp
    timestamp = int(time.time()) + (ttl * 60)

    h = hmac.new(base64.b64decode(key), msg="{0}\n{1}".format(encoded_url, timestamp).encode('utf-8'), digestmod=sha256)
    sas = "SharedAccessSignature sr={0}&sig={1}&se={2}".\
        format(encoded_url, urllib.quote(base64.b64encode(h.digest()), safe=''), timestamp)
    return sas


def post_iot_hub(url, iot_hub, data, sas_token):
    """
        Posts data to Azure IOT hub
    :param url:
    :param data:
    :return:
    """
    headers = {
        "Accept-Encoding": "identity",
        "Host": "{0}.azure-devices.net".format(iot_hub),
        "Content-Type": "application/json",
        "Content-Length": str(getsizeof(data)),
        "Authorization": sas_token
    }
    #logger.debug(headers)
    # r = requests.post(url, data=json.dumps(data), headers=headers, verify=False)
    r = requests.post(url, data=json.dumps(data), headers=headers)
    # print(str(r.status_code))
    if str(r.status_code).startswith('20'):  # success code - 200, 201, 204, etc.
        return True, r.content.decode('utf-8'), r.status_code
    else:
        return False, r.content.decode('utf-8'), r.status_code



@csrf_exempt
@api_view(['GET'])
@permission_classes([AllowAny])
def test_post_iot_hub(request):
    iot_hub = "Askar-output"
    device_id = "ffp-device1"
    primary_key = "neC/RqytrIzBxK2y8aSCKW6rfy2nSTr5VYWn1xja1nE="
    url = constants.IOT_HUB_URL.format(iot_hub, device_id)
    sas_token = generate_sas_token(iot_hub, primary_key, device_id)
    data = {'gw': 123.83219957978865, 'vol': 111111.11111, 'module': 1, 'lat': 33.546385, 'spd': 5, 'id': 'ffp-device1', 'customer': 1, 'temp': 111111.11111, 'dens': 0, 'lon': 71.136510, 't': 'Sep 05,2018 02:18:19 PM', 'type': 3, 'nw': 113.51832126201396}
    status, content, code = post_iot_hub(url, iot_hub, data, sas_token)

    if status:  # post was ok.
        print('Data Posted to Hub')
       # return generic_response((utils.make_response_json(status, content, constants.TEXT_OPERATION_SUCCESSFUL)), http_status=code)
        return generic_response(response_body="Success", http_status=200)

    #else:
       # return utils.generic_response


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
@authentication_classes(())
def temporary_sign_up_hypernet(request):
    response_body = {RESPONSE_MESSAGE: TEXT_PARAMS_MISSING, RESPONSE_STATUS: HTTP_SUCCESS_CODE, RESPONSE_DATA: []}

    User = get_user_model()

    first_name = get_data_param(request, 'first_name', None)
    last_name = get_data_param(request, 'last_name', None)
    password = get_data_param(request, 'password', None)
    email = get_data_param(request, 'email', None)

    enc_password = make_password(password)

    if email and password:

        try:
            user = User()
            user.password = enc_password
            user.email = email
            if first_name:
                user.first_name = first_name
            else:
                user.first_name = email.split('@')[0]

            if last_name:
                user.last_name = last_name
            else:
                user.last_name = email.split('@')[0]
            user.status_id = OptionsEnum.ACTIVE
            user.role_id = RoleTypeEnum.USER
            user.save()

            response_body[RESPONSE_MESSAGE] = "Signed up successfully."
            response_body[RESPONSE_STATUS] = HTTP_SUCCESS_CODE
            response_body[RESPONSE_DATA] = request.data

            return generic_response(response_body, http_status=HTTP_SUCCESS_CODE)

        except:
            traceback.print_exc()
            response_body[RESPONSE_MESSAGE] = "Failed to sign up, try again."
            response_body[RESPONSE_STATUS] = HTTP_ERROR_CODE

            return generic_response(response_body, http_status=HTTP_SUCCESS_CODE)
    else:
        return generic_response(response_body, http_status=HTTP_ERROR_CODE)


@api_view(['GET'])
def evaluate_apis(request):
    import requests
    from openpyxl import load_workbook
    import time
    import xlsxwriter
    filepath = "C:/Users/Taimoor/Desktop/APIs.xlsx"
    result = []
    # load demo.xlsx
    wb = load_workbook(filepath)
    # select demo.xlsx
    sheet = wb.active
    # get max row count
    max_row = sheet.max_row
    # get max column count
    max_column = sheet.max_column
    # iterate over all cells
    # iterate over all rows
    for row in sheet.iter_rows():
        for cell in row:
            if cell.column == 'A':
            # print(cell.value, end=" ")
                if cell.value:
                    start = time.time()
                    response = requests.get('http://' + cell.value,
                                        headers={'Authorization': 'Token a78d7454125bc82d6e6500ddc01feed739106f44'})
                    roundtrip = time.time() - start
                    result.append([cell.value, str(response), str(roundtrip)])

    workbook = xlsxwriter.Workbook('APIs evaluation at ' + str(timezone.now()))
    worksheet = workbook.add_worksheet()

    row = 0
    worksheet.write_row(0, 0, ['API','Response Code', 'Elapsed Time'])

    for col, data in enumerate(result):
        col =0
        worksheet.write_row(row, col, data)
        row +=1
    workbook.close()



def create_location_ent(loc_list):
    for location in loc_list:
        try:
            loc = Entity.objects.get(name=location, type_id=DeviceTypeEntityEnum.LOCATION, status_id= OptionsEnum.ACTIVE)

        except Entity.DoesNotExist:
            loc = Entity()
            loc.name = location
            loc.type_id = DeviceTypeEntityEnum.LOCATION
            loc.status_id = OptionsEnum.ACTIVE
            loc.customer_id = 1
            loc.module_id = 1
            loc.modified_by_id = 1
            loc.save()


def create_contract_location_ass(con_lst, loc_lst):
    for contract, location in zip(con_lst, loc_lst):
        try:
            contract_ent = Entity.objects.get(name=contract, type_id=DeviceTypeEntityEnum.CONTRACT,
                                              status_id=OptionsEnum.ACTIVE, customer_id=1)
        except:
            pass

        try:
            location_ent = Entity.objects.get(name=location, type_id=DeviceTypeEntityEnum.LOCATION,
                                              status_id=OptionsEnum.ACTIVE, customer_id=1)
        except:
            pass


        if contract_ent and location_ent:
            try:
                ass_con_loc = Assignment.objects.get(type_id=DeviceTypeAssignmentEnum.LOCATION_CONTRACT_ASSIGNMENT, status_id=OptionsEnum.ACTIVE,
                                                     parent=location_ent, child=contract_ent)
            except Assignment.DoesNotExist:
                ass_con_loc = Assignment()
                ass_con_loc.name = location_ent.name + "assigned to: "+contract_ent.name
                ass_con_loc.parent = location_ent
                ass_con_loc.child = contract_ent
                ass_con_loc.module_id = ModuleEnum.IOL
                ass_con_loc.customer_id = 1
                ass_con_loc.type_id = DeviceTypeAssignmentEnum.LOCATION_CONTRACT_ASSIGNMENT
                ass_con_loc.status_id = OptionsEnum.ACTIVE
                ass_con_loc.modified_by_id = 1

                ass_con_loc.save()

        if contract_ent:
            bins_assignments = Assignment.objects.filter(child=contract_ent, type_id=DeviceTypeAssignmentEnum.CONTRACT_ASSIGNMENT, status_id=OptionsEnum.ACTIVE).values_list('parent_id', flat=True)
            bins = Entity.objects.filter(id__in=bins_assignments)
            for bin in bins:
                try:
                    bin_location_assignments = Assignment.objects.get(child=location_ent, parent=bin, status_id=OptionsEnum.ACTIVE,
                                                                      type_id=DeviceTypeAssignmentEnum.LOCATION_ASSIGNMENT)
                except Assignment.DoesNotExist:
                    ass_con_loc = Assignment()
                    ass_con_loc.name = location_ent.name + "assigned to: " + bin.name
                    ass_con_loc.parent = bin
                    ass_con_loc.child = location_ent
                    ass_con_loc.module_id = ModuleEnum.IOL
                    ass_con_loc.customer_id = 1
                    ass_con_loc.type_id = DeviceTypeAssignmentEnum.LOCATION_ASSIGNMENT
                    ass_con_loc.status_id = OptionsEnum.ACTIVE
                    ass_con_loc.modified_by_id = 1

                    ass_con_loc.save()





def save_contract_location_ass(request=None):
    import xlrd
    fname = 'ioa/tests/Contracts_Nov 29 2018 (1).xlsx'

    xl_workbook = xlrd.open_workbook(fname)

    xl_sheet = xl_workbook.sheet_by_index(0)

    xl_sheet.cell_value(1, 0)


    contract_lst = []
    locations_lst = []

    for i in range(1, xl_sheet.nrows):
        try:
            if xl_sheet.cell_value(i, 6) is not '':
                val_locations = xl_sheet.cell_value(i, 6)
                val_contracts = int(xl_sheet.cell_value(i, 0))
                contract_lst.append(val_contracts)
                locations_lst.append(val_locations)

            else:
                continue
        except:
            continue


    print(len(contract_lst))
    print(len(locations_lst))
    #create_locations = create_location_ent(loc_list=locations_lst)

    create_contract_location_ass(con_lst=contract_lst, loc_lst=locations_lst)




def create_ent_with_type(lst, type_id, start_datetime = None, end_datetime = None, client_lst=None):
    for ent, start_time, end_time, client in zip(lst, start_datetime, end_datetime, client_lst):
        try:
            loc = Entity.objects.get(name=ent, type_id=type_id, status_id= OptionsEnum.ACTIVE)

        except Entity.DoesNotExist:
            loc = Entity()
            loc.name = ent
            loc.type_id = type_id
            loc.status_id = OptionsEnum.ACTIVE
            loc.customer_id = 2
            loc.module_id = 1
            loc.modified_by_id = 17

            if type_id == DeviceTypeEntityEnum.CONTRACT:
                try:
                    client = CustomerClients.objects.get(name=client)
                    loc.client=client
                    loc.date_commissioned = start_time
                    loc.date_of_joining = end_time
                except:
                    loc.client=None
            loc.save()



def create_clients(clients_list, part_code_lst):
    for cl, ptc in zip(clients_list, part_code_lst):
        try:
            client = CustomerClients.objects.get(name=cl, status_id= OptionsEnum.ACTIVE)

        except CustomerClients.DoesNotExist:
            print(type(cl))
            client = CustomerClients()
            client.name = cl
            client.customer_id = 2
            client.party_code = ptc
            client.modified_by_id = 17
            client.status_id = OptionsEnum.ACTIVE
            client.save()


def create_area_contract_loc_assignment(contract_lst, area_lst, loc_lst):
    for contract, area in zip(contract_lst, area_lst):
        try:
            contract_ent = Entity.objects.get(name=contract, type_id=DeviceTypeEntityEnum.CONTRACT,
                                              status_id=OptionsEnum.ACTIVE, customer_id=2)
        except:
            contract_ent = None
            pass

        try:
            area_ent = Entity.objects.get(name=area, type_id=DeviceTypeEntityEnum.AREA,
                                              status_id=OptionsEnum.ACTIVE, customer_id=2)
        except:
            area_ent = None
            pass


        if contract_ent and area_ent:
            try:
                ass_area_contract = Assignment.objects.get(type_id=DeviceTypeAssignmentEnum.AREA_CONTRACT_ASSIGNMENT, status_id=OptionsEnum.ACTIVE,
                                                     parent=area_ent, child=contract_ent, customer_id=2)
            except Assignment.DoesNotExist:
                ass_area_contract = Assignment()
                ass_area_contract.name = contract_ent.name + "assigned to: "+area_ent.name
                ass_area_contract.parent = area_ent
                ass_area_contract.child = contract_ent
                ass_area_contract.module_id = ModuleEnum.IOL
                ass_area_contract.customer_id = 2
                ass_area_contract.type_id = DeviceTypeAssignmentEnum.AREA_CONTRACT_ASSIGNMENT
                ass_area_contract.status_id = OptionsEnum.ACTIVE
                ass_area_contract.modified_by_id = 17

                ass_area_contract.save()


    for contract, location in zip(contract_lst, loc_lst):
        try:
            contract_ent = Entity.objects.get(name=contract, type_id=DeviceTypeEntityEnum.CONTRACT,
                                              status_id=OptionsEnum.ACTIVE, customer_id=2)
        except:
            contract_ent = None
            pass

        try:
            location_ent = Entity.objects.get(name=location, type_id=DeviceTypeEntityEnum.LOCATION,
                                              status_id=OptionsEnum.ACTIVE, customer_id=2)
        except:
            location_ent = None
            pass


        if contract_ent and location_ent:
            try:
                ass_con_loc = Assignment.objects.get(type_id=DeviceTypeAssignmentEnum.LOCATION_CONTRACT_ASSIGNMENT, status_id=OptionsEnum.ACTIVE,
                                                     parent=location_ent, child=contract_ent, customer_id=2)
            except Assignment.DoesNotExist:
                ass_con_loc = Assignment()
                ass_con_loc.name = location_ent.name + "assigned to: "+contract_ent.name
                ass_con_loc.parent = location_ent
                ass_con_loc.child = contract_ent
                ass_con_loc.module_id = ModuleEnum.IOL
                ass_con_loc.customer_id = 2
                ass_con_loc.type_id = DeviceTypeAssignmentEnum.LOCATION_CONTRACT_ASSIGNMENT
                ass_con_loc.status_id = OptionsEnum.ACTIVE
                ass_con_loc.modified_by_id = 17

                ass_con_loc.save()




def create_generic_ent(lst,type_id):
    for ent in lst:
        try:
            loc = Entity.objects.get(name=ent, type_id=type_id, status_id= OptionsEnum.ACTIVE)

        except Entity.DoesNotExist:
            loc = Entity()
            loc.name = ent
            loc.type_id = type_id
            loc.status_id = OptionsEnum.ACTIVE
            loc.customer_id = 2
            loc.module_id = 1
            loc.modified_by_id = 17
            loc.save()


@api_view(['POST'])
@exception_handler(generic_response(response_body=ERROR_RESPONSE_BODY, http_status=HTTP_ERROR_CODE))
def upload_data_cms(request=None):
    files = request.FILES.getlist('files')
    import xlrd
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile

    response = {RESPONSE_MESSAGE: "Successful"}
    response[RESPONSE_DATA] = {TEXT_OPERATION_SUCCESSFUL: True}
    response[RESPONSE_STATUS] = HTTP_SUCCESS_CODE
    for f in files:
        fname = default_storage.save(f.name, ContentFile(f.read()))
        xl_workbook = xlrd.open_workbook('media/'+fname)
        xl_sheet = xl_workbook.sheet_by_index(0)
        xl_sheet.cell_value(1, 0)

        for i in range(1, xl_sheet.nrows):
            try:
                obj = dict()
                obj['customer'] = 13
                obj['timestamp'] = parse_datefield(xl_sheet.cell_value(i, 0))
                #create entities of type truck
                obj['vehicle'] = get_or_create_entity(xl_sheet.cell_value(i, 2), DeviceTypeEntityEnum.TRUCK)
                truck = obj['vehicle']
                obj['vehicle'] = truck.id
                truck, obj['office'] = truck_office_wheels(truck, xl_sheet.cell_value(i, 65))
                truck = truck_make_model(truck, xl_sheet.cell_value(i, 66))
                truck.save()
                obj['loading_location'] = get_or_create_entity(xl_sheet.cell_value(i, 3), DeviceTypeEntityEnum.LOCATION).id
                obj['loading_city'] = get_or_create_entity(xl_sheet.cell_value(i, 4), DeviceTypeEntityEnum.AREA).id
                obj['destination'] = get_or_create_entity(xl_sheet.cell_value(i, 5), DeviceTypeEntityEnum.AREA).id
                obj['vms'] = xl_sheet.cell_value(i, 6)
                obj['trip_number'] = xl_sheet.cell_value(i, 7)
                obj['order_number'] = xl_sheet.cell_value(i, 8)
                obj['client'] = get_or_create_client(xl_sheet.cell_value(i, 9))
                obj['trip_start_datetime'] = parse_datefield(xl_sheet.cell_value(i, 10))
                obj['loaded_datetime'] = parse_datefield(xl_sheet.cell_value(i, 11))
                obj['stops_loaded_duration'] = parse_timefield(xl_sheet.cell_value(i, 12))
                obj['arrival_datetime'] = parse_datefield(xl_sheet.cell_value(i, 13))
                obj['unloaded_datetime'] = parse_datefield(xl_sheet.cell_value(i, 14))
                obj['stops_unloading_duration'] = parse_timefield(xl_sheet.cell_value(i, 15))
                obj['arrival_datetime'] = parse_datefield(xl_sheet.cell_value(i, 16))

                obj['halting'] = parse_intfield(xl_sheet.cell_value(i, 20))
                # regionl_closing_month = parse_datefield(xl_sheet.cell_value(i, 21))


                obj['loaded_workshop_in'] = parse_datefield(xl_sheet.cell_value(i, 53))
                obj['loaded_workshop_out'] = parse_datefield(xl_sheet.cell_value(i, 54))
                #calculate duration
                obj['loaded_work_order_number'] = xl_sheet.cell_value(i, 55)
                obj['loaded_workshop_remarks'] = xl_sheet.cell_value(i, 56)

                obj['unloaded_workshop_in'] = parse_datefield(xl_sheet.cell_value(i, 57))
                obj['unloaded_workshop_out'] = parse_datefield(xl_sheet.cell_value(i, 58))
                obj['unloaded_work_order_no'] = xl_sheet.cell_value(i, 59)
                obj['unloaded_workshop_remarks'] = xl_sheet.cell_value(i, 60)

                obj['km_loaded'] = parse_intfield(xl_sheet.cell_value(i, 62))
                obj['km_unloaded'] = parse_intfield(xl_sheet.cell_value(i, 63))
                obj['total_km'] = parse_intfield(xl_sheet.cell_value(i, 64))


                obj['supervisor'] = get_or_create_entity(xl_sheet.cell_value(i, 69), DeviceTypeEntityEnum.SUPERVISOR).id

                serializer = CMSVehicleReportingSerializer(data=obj, context={'request': request})
                if serializer.is_valid():
                    entity = serializer.save()
                else:
                    print(serializer.errors)
            except:
                print('Serial Number of erroneous record: '+str(xl_sheet.cell_value(i, 1)))
                traceback.print_exc()
        return generic_response(response_body=response, http_status=200)


def parse_datefield(value):
    import xlrd
    from datetime import datetime
    if value:
        return datetime(*xlrd.xldate_as_tuple(value, 0))
    else:
        return None


def parse_timefield(value):
    import xlrd
    from datetime import datetime
    if value:
        val = xlrd.xldate_as_tuple(value, 0)
        hour = int(val[3])
        minute = int(val[4])
        return hour*60 + minute
    else:
        return None


def parse_intfield(value):
    if value:
        return int(value)
    else:
        return 0


def get_or_create_entity(entity_name, type):

    obj, created = Entity.objects.get_or_create(name=entity_name, type_id=type,
                                                customer_id=13,
                                                module_id=1,
                                                status_id=OptionsEnum.ACTIVE,
                                                modified_by_id=1)
    return obj


def get_or_create_client(client_name):

    obj, created = CustomerClients.objects.get_or_create(name=client_name,
                                                customer_id=13,
                                                status_id=OptionsEnum.ACTIVE,
                                                modified_by_id=1)

    return obj.id


def truck_make_model(truck, row):
    data = row.split()
    truck.make = data[0]
    truck.model = data[1]
    return truck


def truck_office_wheels(truck, row):
    data = row.split('-')
    office = data[0]
    data2 = data[1].split()
    truck.wheels = data2[0]
    return truck, office


def clean_data_duplicate_bins(request=None):
    try:
        all_bins = Entity.objects.filter(customer_id=2, type_id=21)
        print('All bins '+ str(all_bins.count()))
        unique_bins = all_bins.values_list('name', flat=True).distinct()
        print('Unique '+ str(unique_bins.count()))
        for b in unique_bins:
            dupes = all_bins.filter(name=b).order_by('-created_datetime')
            if dupes.count() >1:

                latest_bin = dupes[0]
                print('count of dupes:'+str(dupes.count()))
                print('Bin:'+latest_bin.name)
                # set a flag in con so we dont show this in list no more
                con = Assignment.objects.get(parent=latest_bin, type_id=DeviceTypeAssignmentEnum.CONTRACT_ASSIGNMENT).child
                print('Contract:'+con.name)
                con.volume = True
                con.save()
                to_delete = dupes.exclude(id=latest_bin.id)
                print('     latest bin: '+str(latest_bin.id) + 'created:' + str(latest_bin.created_datetime))
                for d in to_delete:
                    print('         to delete: '+str(d.id)+ 'created:' + str(d.created_datetime))
                to_delete.delete()

    except Exception as e:
        traceback.print_exc()







def create_cms_entities(request=None):
    filepath = 'ioa/tests/devices.txt'
    devices = []
    ents = []
    with open(filepath) as fp:
        line = fp.readline()
        cnt = 1
        while line:
            temp = line.split(',')
            devices.append('cms-truck-' + temp[0].strip())
            ents.append(temp[1].strip())

            line = fp.readline()
            cnt += 1

    for device,ent in zip(devices, ents):

       # c_d = create_customer_device(device)
       # c_d.save()

        try:
            entity = Entity.objects.get(name=ent)

            c_d = CustomerDevice.objects.get(device_id = device)
            entity.device_name = c_d
            entity.save()
        except:
            ent = Entity(
                name=ent,
                type_id=DeviceTypeEntityEnum.TRUCK,
                status_id=OptionsEnum.ACTIVE,
                module_id=ModuleEnum.IOL,
                customer_id=13,
                modified_by_id=334,
                device_name_id=c_d.id,
                entity_sub_type_id=IOFOptionsEnum.WHEELER_6
            )
            ent.save()



def change_customer_devices(request=None):

    ents =Entity.objects.filter(customer_id=13)

    for ent in ents:
        customer_device = ent.device_name
        customer_device.device_id = 'cms-truck-'+ent.name
        customer_device.save()


def create_customer_device(device_id):

    try:
        c_d = CustomerDevice.objects.get(device_id=device_id)
    except:
        c_d = CustomerDevice(
            device_id = device_id,
            status_id = OptionsEnum.ACTIVE,
            customer_id = 13,
            type_id = DeviceTypeEntityEnum.TRUCK,
            module_id = ModuleEnum.IOL,
            connection_string='askar-output',
            primary_key = get_random_string(length=9, allowed_chars='abcdefghijklmono123456'),
        )

    return c_d




def create_new_customer_device(request=None):

    ents = Entity.objects.filter(customer_id=13, type_id=3)

    for ent in ents:
        if ent.device_name:
            pass
        else:
            name = ent.name
            name = name.lower()
            name =name.replace(" ","")
            device = create_customer_device('cms-truck-'+name)
            device.save()
            ent.device_name = device
            ent.save()

